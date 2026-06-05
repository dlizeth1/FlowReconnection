# -*- coding: utf-8 -*-
"""
live_sheet.py — Lee EN VIVO el .xlsx de Flow Reconnection desde Google Drive y
devuelve los supuestos del modelo, para que config.py los sobrescriba ("Sheet manda").

FUENTE: Google Drive file 13v_c-IKpTtp25rWFLb47jb0oLEccQ7Fi
        ("Financiero - centro flotación.xlsx"), descargado por link público.

Diseño:
  - Descarga el binario .xlsx vía la URL pública de Drive (uc?export=download).
  - Parsea con openpyxl POR ETIQUETA (no por celda fija): tolera que se inserten
    o muevan filas en el Sheet sin romper el mapeo.
  - Cada sección se extrae de forma defensiva: si una etiqueta cambió, esa sección
    se omite (queda el valor estático de config.py) en vez de tumbar todo.

Requisitos: requests, openpyxl  (ver requirements.txt)

Uso directo (diagnóstico):
    python live_sheet.py          # imprime lo que leyó del Sheet
"""

import io
import re
import sys
import unicodedata

FILE_ID = "13v_c-IKpTtp25rWFLb47jb0oLEccQ7Fi"
DOWNLOAD_URL = f"https://drive.google.com/uc?export=download&id={FILE_ID}"
DEFAULT_TIMEOUT = 60


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def _norm(s):
    """Normaliza texto para comparar etiquetas: sin acentos, minúsculas, sin espacios extra."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s).strip().lower()


def _num(v):
    """Coacciona a número valores como 43610400, '$150.000', '14 horas', '4,000 usd'."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v)
    s = s.replace("$", "").replace("usd", "").replace("USD", "")
    # quita separadores de miles (coma o punto seguido de exactamente 3 dígitos)
    s = re.sub(r"[.,](?=\d{3}\b)", "", s)
    m = re.search(r"-?\d+(?:[.,]\d+)?", s)
    if not m:
        return None
    token = m.group(0).replace(",", ".")
    try:
        f = float(token)
        return int(f) if f.is_integer() else f
    except ValueError:
        return None


def _sheet(wb, name):
    """Devuelve la hoja cuyo nombre coincide (insensible a acentos/caso)."""
    target = _norm(name)
    for title in wb.sheetnames:
        if _norm(title) == target:
            return wb[title]
    # match parcial como respaldo
    for title in wb.sheetnames:
        if target in _norm(title):
            return wb[title]
    raise KeyError(f"hoja no encontrada: {name!r} (hojas: {wb.sheetnames})")


def _find_cell(ws, label, max_col=14, max_row=80, exact=False):
    """Busca la primera celda cuyo texto coincide con `label`. Devuelve (row, col) o None."""
    target = _norm(label)
    for r in range(1, min(ws.max_row, max_row) + 1):
        for c in range(1, min(ws.max_column, max_col) + 1):
            cell = _norm(ws.cell(r, c).value)
            if not cell:
                continue
            if (cell == target) if exact else (target in cell):
                return (r, c)
    return None


def _row_of(ws, label, **kw):
    hit = _find_cell(ws, label, **kw)
    return hit[0] if hit else None


# ---------------------------------------------------------------------------
# Descarga
# ---------------------------------------------------------------------------
def fetch_workbook(timeout=DEFAULT_TIMEOUT):
    import requests
    import openpyxl

    r = requests.get(DOWNLOAD_URL, allow_redirects=True, timeout=timeout)
    r.raise_for_status()
    if not r.content[:2] == b"PK":
        raise ValueError(
            "la descarga no es un .xlsx (¿el archivo dejó de ser público "
            "o Drive devolvió una página de aviso?)"
        )
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)


# ---------------------------------------------------------------------------
# Extracción por sección (cada una defensiva)
# ---------------------------------------------------------------------------
def _capex(wb, ov, warn):
    ws = _sheet(wb, "INVERSIÓN")
    # columna D (4) = TOTAL por concepto; etiquetas en columna A.
    items = [
        ("tanques_flotacion", "tanque de flotacion"),
        ("adecuaciones_y_muebles", "adecuaciones y muebles"),
        ("capital_trabajo_6m", "capital de trabajo"),
        ("pagina_web", "pagina web"),
        ("seguridad_camaras", "seguridad - camaras"),
        ("creacion_marca", "creacion de marca"),
        ("insumos", "insumos"),
    ]
    capex = {}
    for key, label in items:
        row = _row_of(ws, label, max_col=2)
        if row:
            val = _num(ws.cell(row, 4).value)
            if val is not None and val != 0:
                capex[key] = int(val)
    if capex:
        ov["CAPEX"] = capex
        # total del Sheet (fila TOTAL, col D) si existe; si no, suma de items.
        trow = _row_of(ws, "total", max_col=2)
        total = _num(ws.cell(trow, 4).value) if trow else None
        ov["CAPEX_TOTAL"] = int(total) if total else int(sum(capex.values()))
    # referencias del tanque
    trow = _row_of(ws, "tanque de flotacion", max_col=2)
    if trow:
        unit = _num(ws.cell(trow, 3).value)        # C: valor unitario
        rent = _num(ws.cell(trow, 6).value)        # F: renting total
        dep = _num(ws.cell(trow, 7).value)         # G: depreciación mensual
        if unit:
            ov["TANQUE_VALOR_UNITARIO"] = int(unit)
        if rent:
            ov["TANQUE_RENTING_TOTAL"] = int(rent)
        if dep:
            ov["DEPRECIACION_MENSUAL"] = int(dep)


def _capacidad(wb, ov, warn):
    ws = _sheet(wb, "CAPACIDAD INSTALADA")
    cap = {}
    mapa = [
        ("horas_por_flote", "tiempo por flote"),
        ("sesiones_dia_por_tanque", "sesiones x dia x tanque"),
        ("dias_abiertos_mes", "dias abiertos x mes"),
    ]
    for key, label in mapa:
        row = _row_of(ws, label, max_col=1)
        if row:
            val = _num(ws.cell(row, 2).value)   # col B = 1 tanque
            if val is not None:
                cap[key] = val
    row = _row_of(ws, "horas de atencion", max_col=1)
    if row:
        val = _num(ws.cell(row, 2).value)       # "14 horas" -> 14
        if val is not None:
            cap["horas_atencion_dia"] = int(val)
    if cap:
        ov["CAPACIDAD"] = cap
    row = _row_of(ws, "valor tentativo del servicio", max_col=1)
    if row:
        val = _num(ws.cell(row, 2).value)
        if val:
            ov["PRECIO_OPERATIVO_PLANO"] = int(val)


def _precios_franja(wb, ov, warn):
    ws = _sheet(wb, "INGRESOS")
    # tabla "Ingresos mensuales ... (ocupación al 100%)": filas Diurnas/Nocturnas
    drow = _row_of(ws, "diurnas", max_col=1)
    nrow = _row_of(ws, "nocturnas", max_col=1)
    # shares en col D de "57% de sesiones diurnas..." / "43% de sesiones nocturnas..."
    srow_d = _row_of(ws, "sesiones diurnas", max_col=1)
    srow_n = _row_of(ws, "sesiones nocturnas", max_col=1)
    franja = {}
    if drow:
        pd = _num(ws.cell(drow, 3).value)       # col C = valor unitario
        sd = _num(ws.cell(srow_d, 4).value) if srow_d else 0.57
        if pd:
            franja["diurno"] = {"precio": int(pd), "share": float(sd), "franja": "9am-5pm"}
    if nrow:
        pn = _num(ws.cell(nrow, 3).value)
        sn = _num(ws.cell(srow_n, 4).value) if srow_n else 0.43
        if pn:
            franja["nocturno"] = {"precio": int(pn), "share": float(sn), "franja": "5pm-11pm"}
    if len(franja) == 2:
        ov["PRECIOS_FRANJA"] = franja


def _costos_variables(wb, ov, warn):
    # Total de insumos (costos variables a 100% capacidad, 4 tanques) = INVERSIÓN!Insumos
    ws = _sheet(wb, "INVERSIÓN")
    row = _row_of(ws, "insumos", max_col=2)
    if row:
        val = _num(ws.cell(row, 4).value)
        if val:
            ov["COSTOS_VARIABLES_100_TOTAL"] = int(val)


# Mapeo clave config -> etiqueta de fila en el P&L MEDELLÍN
_FIJOS_LABELS = [
    ("arriendo", "arriendo mensual"),
    ("servicios_publicos", "servicios publicos"),
    ("servicios_generales", "servicios generales"),
    ("administracion", "administracion"),
    ("recepcion", "recepcion"),
    ("contabilidad", "contabilidad"),
    ("software", "software"),
    ("seguridad_camaras", "seguridad - camaras - alarma"),
    ("seguro", "seguro"),
    ("pagina_web_mant", "pagina web - mantenimiendo"),
    ("pauta_anuncios", "inversion en pauta"),
    ("mercadeo", "mercadeo"),
    ("mantenimiento", "mantenimiento"),
]


def _meses_cols(ws):
    """Columnas de los meses (desde B hasta antes del total anual)."""
    # fila de meses: la que contiene 'JULIO' o 'AGOSTO'
    mrow = _row_of(ws, "julio", max_col=14) or _row_of(ws, "agosto", max_col=14)
    cols = []
    if mrow:
        for c in range(2, 12):
            if _norm(ws.cell(mrow, c).value):
                cols.append(c)
    if not cols:
        cols = list(range(2, 11))   # respaldo: B..J
    return cols


def _costos_fijos(wb, ov, warn):
    ws = _sheet(wb, "MEDELLÍN")
    cols = _meses_cols(ws)
    steady_col = cols[-1]           # último mes = estado estable
    fijos = {}
    for key, label in _FIJOS_LABELS:
        row = _row_of(ws, label, max_col=1)
        if row:
            val = _num(ws.cell(row, steady_col).value)
            fijos[key] = int(val) if val is not None else 0
    if len(fijos) >= 10:
        ov["COSTOS_FIJOS_4T"] = fijos
    # Depreciación mensual: fila DEPRECIACIÓN del P&L (lugar estable),
    # no INVERSIÓN (esa columna puede desaparecer del Sheet).
    drow = _row_of(ws, "depreciacion", max_col=1)
    if drow:
        dep = _num(ws.cell(drow, steady_col).value)
        if dep:
            ov["DEPRECIACION_MENSUAL"] = int(dep)


def _ocupacion(wb, ov, warn):
    ws = _sheet(wb, "MEDELLÍN")
    cols = _meses_cols(ws)
    row = _row_of(ws, "% ocupacion", max_col=1)
    if row:
        curva = []
        for c in cols:
            v = _num(ws.cell(row, c).value)
            if v is not None:
                curva.append(round(float(v), 4))
        if curva:
            # el modelo usa 12 meses; rellena con el último valor estable
            while len(curva) < 12:
                curva.append(curva[-1])
            ov["OCUPACION_OBJETIVO"] = curva


def _escenario(ws):
    """Extrae un escenario reportado (Medellín/Barranquilla) por etiquetas."""
    cols = _meses_cols(ws)
    total_col = cols[-1] + 1        # columna del total anual (justo después del último mes)
    esc = {}
    # sesiones
    srow = _row_of(ws, "servicios al mes", max_col=1)
    if srow:
        esc["sesiones"] = [int(_num(ws.cell(srow, c).value) or 0) for c in cols]
    # ingresos total
    irow = _row_of(ws, "ingresos", max_col=1)
    if irow:
        t = _num(ws.cell(irow, total_col).value)
        esc["ingresos_total"] = int(t) if t else int(sum(_num(ws.cell(irow, c).value) or 0 for c in cols))
    # total costos
    crow = _row_of(ws, "total costos", max_col=1)
    if crow:
        t = _num(ws.cell(crow, total_col).value)
        if t is not None:
            esc["costos_total"] = int(t)
    # utilidad marginal
    urow = _row_of(ws, "utilidad marginal", max_col=1)
    if urow:
        t = _num(ws.cell(urow, total_col).value)
        if t is not None:
            esc["utilidad_marginal"] = int(t)
    # después de depreciación = fila siguiente a DEPRECIACIÓN, columna total
    drow = _row_of(ws, "depreciacion", max_col=1)
    if drow:
        t = _num(ws.cell(drow + 1, total_col).value)
        if t is not None:
            esc["despues_depreciacion"] = int(t)
    return esc


def _escenarios(wb, ov, warn):
    out = {}
    try:
        m = _escenario(_sheet(wb, "MEDELLÍN"))
        if m.get("ingresos_total"):
            m["precio"] = 155_000
            out["Medellín (4 tanques)"] = m
    except KeyError:
        pass
    try:
        b = _escenario(_sheet(wb, "BARRANQUILLA"))
        if b.get("ingresos_total"):
            b["precio"] = 155_000
            out["Barranquilla (2–3 tanques)"] = b
    except KeyError:
        pass
    if out:
        ov["ESCENARIOS_REPORTADOS"] = out


# NOTA: el cap table (hoja ABONOS) NO se extrae a propósito: contiene datos
# personales (nombres y cédulas) que no deben llegar al HTML público.
_SECTIONS = [_capex, _capacidad, _precios_franja, _costos_variables,
             _costos_fijos, _ocupacion, _escenarios]


def extract(wb, verbose=False):
    """Devuelve dict de overrides extraídos del workbook. Cada sección es defensiva."""
    ov = {}
    warnings = []
    for fn in _SECTIONS:
        try:
            fn(wb, ov, warnings.append)
        except Exception as e:   # una sección rota no debe tumbar el resto
            warnings.append(f"{fn.__name__}: {e}")
    if verbose and warnings:
        for w in warnings:
            print(f"[live_sheet] aviso: {w}", file=sys.stderr)
    return ov


def fetch_overrides(timeout=DEFAULT_TIMEOUT, verbose=False):
    """Descarga el Sheet y devuelve los overrides. Lanza excepción si falla la descarga."""
    wb = fetch_workbook(timeout=timeout)
    return extract(wb, verbose=verbose)


def apply_live_overrides(g, timeout=DEFAULT_TIMEOUT, verbose=False):
    """
    Aplica los valores del Sheet sobre el namespace `g` (globals() de config).
    'Sheet manda': sobrescribe y recalcula los totales derivados.
    Devuelve el dict de overrides aplicados (vacío si no aplicó nada).
    """
    ov = fetch_overrides(timeout=timeout, verbose=verbose)

    # Overrides directos
    for key in ("CAPEX", "CAPEX_TOTAL", "TANQUE_VALOR_UNITARIO", "TANQUE_RENTING_TOTAL",
                "DEPRECIACION_MENSUAL", "PRECIO_OPERATIVO_PLANO", "PRECIOS_FRANJA",
                "COSTOS_VARIABLES_100_TOTAL", "COSTOS_FIJOS_4T", "OCUPACION_OBJETIVO",
                "ESCENARIOS_REPORTADOS"):
        if key in ov:
            if key == "CAPACIDAD":
                continue
            g[key] = ov[key]

    # CAPACIDAD: merge (conserva claves no presentes en el Sheet)
    if "CAPACIDAD" in ov and isinstance(g.get("CAPACIDAD"), dict):
        g["CAPACIDAD"].update(ov["CAPACIDAD"])

    # Recalcular totales derivados
    if "COSTOS_FIJOS_4T" in ov:
        g["COSTOS_FIJOS_4T_TOTAL"] = sum(g["COSTOS_FIJOS_4T"].values())
    if "CAPEX_TOTAL" in ov:
        # la "Base" del slider y el default siguen al capex vivo
        if isinstance(g.get("CAPEX_ALTERNATIVAS"), dict):
            g["CAPEX_ALTERNATIVAS"]["Base (Sheet en vivo)"] = g["CAPEX_TOTAL"]
        if isinstance(g.get("DEFAULTS"), dict):
            g["DEFAULTS"]["capex"] = g["CAPEX_TOTAL"]

    g["LIVE_DATA"] = bool(ov)
    g["LIVE_SOURCE"] = DOWNLOAD_URL
    return ov


if __name__ == "__main__":
    import json
    ov = fetch_overrides(verbose=True)
    print(f"Secciones leídas del Sheet: {len(ov)}\n")
    print(json.dumps(ov, ensure_ascii=False, indent=2, default=str))
